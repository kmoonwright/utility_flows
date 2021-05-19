import openpyxl

xlsx = openpyxl.load_workbook('roster_X334.xlsx')

## getting the sheet to active
sheet = xlsx.active

# SINGLE CELL DATA
name = sheet['A1']
tag = sheet.cell(row=1, column=2)
## printing the values of cells
# print(name.value)
# print(tag.value)

# GET COLUMN DATA
## getting the reference of the cells which we want to get the data from
columns = sheet.columns
# for column in columns:
    # for cell in column:
        # print(cell.value, end=' ')
    # print("\n")


# GET ROW DATA
rows = sheet.rows
## printing the values of cells using rowsipython
# for row in rows:
#     for cell in row:
#         print(cell.value, end = ' ')
#         break

#     print("\n")

qualification_courses = [
    "Second-Language Acquisition",
    "Grammar Fundamentals for ESL Teachers",
    "Methods and Materials for Teaching English as a Second Language",
    "Cross-Cultural Communication",
    "Teaching Pronunciation as a Communicative Skill",
    "Fundamentals of Linguistics for ESL Teachers",
    "Calculus 1"
]

roster = {}
# Create dict for - email: [courses]
for x in range (6, 511):
    email = sheet.cell(row=x,column=5).value
    course = sheet.cell(row=x,column=8).value
    if email in roster.keys():
        # add to val
        roster[email].append(course)
    else:
        roster[email] = [course]

# print(f"Roster: {roster}\n")

# Check against qualifications
qualified, unqualified = [], []
for student in roster:
    course_list = roster[student]
    qual_course_count = 0
    for course in course_list:
        if course == "Practicum in Course Design for ESL/EFL Teachers":
            unqualified.append(student)
            break
        if course in qualification_courses:
            qual_course_count += 1

    if qual_course_count >= 4:
        qualified.append(student)
    else:
        unqualified.append(student)

print(f"Qualified students: {qualified}\n")
print(f"Unqalified students: {unqualified}")

for student in qualified:
    print(f'{student}\n')

# for email, course_list in roster.items():
# 	print(email)
# 	for courses in course_list:
# 		print(courses)